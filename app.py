import streamlit as st
from openai import OpenAI
from dotenv import load_dotenv
import os
import json
import io
import base64
from datetime import datetime
from pathlib import Path

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import docx
except ImportError:
    docx = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

import csv

# Load environment variables
load_dotenv()

# Initialize OpenAI client
@st.cache_resource
def get_openai_client():
    return OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

client = get_openai_client()

# Page config
st.set_page_config(
    page_title="BIM-CHATBOT",
    page_icon="🔱",
    layout="wide"
)

def get_base64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except FileNotFoundError:
        return ""

flag_base64 = get_base64("assets/barbados_flag.png")
trident_base64 = get_base64("assets/trident background.png")
brand_base64 = flag_base64 or trident_base64

# BIM-CHATBOT Gold Theme
trident_bg = f'url("data:image/png;base64,{trident_base64}")' if trident_base64 else 'none'
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

* {{
    box-sizing: border-box;
    font-family: 'Inter', Arial, sans-serif;
}}

html, body, [class*="css"] {{
    background: #f1ece6;
}}

[data-testid="stAppViewContainer"] {{
    background: #f1ece6;
}}

[data-testid="stMain"] {{
    position: relative;
    margin: 24px 20px 20px 0;
    border-radius: 0 16px 16px 0;
    border: 1px solid rgba(0, 0, 0, 0.08);
    background: linear-gradient(180deg, #f7f3ef 0%, #f1ece6 100%) !important;
    overflow: hidden;
    box-shadow: 0 18px 35px rgba(0, 0, 0, 0.08);
}}

[data-testid="stMain"]::before {{
    content: "";
    position: absolute;
    inset: 0;
    background-image: {trident_bg};
    background-repeat: no-repeat;
    background-position: center 60%;
    background-size: min(70vw, 700px);
    opacity: 0.08;
    filter: grayscale(1) brightness(1.35);
    pointer-events: none;
    z-index: 0;
}}

[data-testid="stMain"] > div {{
    position: relative;
    z-index: 1;
    padding-top: 1rem;
}}

.block-container {{
    max-width: 1080px;
    padding-top: 1.2rem;
    padding-left: 1.8rem;
    padding-right: 1.8rem;
    padding-bottom: 1rem;
}}

header[data-testid="stHeader"] {{
    background: #f3c614 !important;
}}

[data-testid="stSidebar"] {{
    background: radial-gradient(120% 120% at 10% 10%, #06195f 0%, #020a2a 45%, #01051a 100%) !important;
    border-right: 1px solid rgba(255, 255, 255, 0.12) !important;
}}

[data-testid="stSidebar"] * {{
    color: #edf0ff !important;
}}

.sidebar-brand {{
    display: flex;
    align-items: center;
    gap: 12px;
    margin-top: 0.4rem;
    margin-bottom: 1.1rem;
}}

.sidebar-brand img {{
    width: 46px;
    height: 36px;
    object-fit: cover;
    border-radius: 6px;
    box-shadow: 0 6px 16px rgba(0, 0, 0, 0.35);
}}

.sidebar-brand span {{
    color: #eecf77 !important;
    font-size: 1.9rem;
    font-weight: 800;
    letter-spacing: 0.4px;
}}

[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"],
[data-testid="stSidebar"] [data-testid="stTextInput"] input,
[data-testid="stSidebar"] .stButton > button {{
    background: rgba(255, 255, 255, 0.05) !important;
    border: 1px solid rgba(255, 255, 255, 0.14) !important;
    border-radius: 14px !important;
    color: #f8f9ff !important;
    min-height: 48px;
}}

/* Force black text for the AI model selector values (e.g., GPT-* labels). */
[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] span,
[data-testid="stSidebar"] [data-testid="stSelectbox"] div[data-baseweb="select"] * {{
    color: #000000 !important;
}}

/* Force black text for file uploader instructions/help text in the sidebar. */
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] *,
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzoneInstructions"] * {{
    color: #000000 !important;
}}

[data-testid="stSidebar"] .stButton > button {{
    font-weight: 600;
    box-shadow: 0 8px 22px rgba(0, 0, 0, 0.2);
}}

[data-testid="stSidebar"] .stButton > button:hover {{
    border-color: rgba(243, 198, 20, 0.6) !important;
    color: #f3c614 !important;
}}

.top-strip {{
    display: flex;
    justify-content: flex-end;
    align-items: center;
    background: #f3c614;
    border-radius: 10px;
    padding: 8px 12px;
    margin: -0.25rem -0.6rem 1.2rem -0.6rem;
}}

.top-strip .deploy-btn {{
    border: 1px solid rgba(0, 0, 0, 0.18);
    border-radius: 12px;
    padding: 5px 16px;
    font-size: 1rem;
    color: #111827;
    background: rgba(255, 211, 40, 0.9);
    font-weight: 500;
}}

.main-title-wrap {{
    display: flex;
    align-items: center;
    gap: 16px;
    margin-bottom: 6px;
}}

.main-title-wrap img {{
    width: 58px;
    height: 44px;
    object-fit: cover;
    border-radius: 8px;
    box-shadow: 0 6px 16px rgba(0, 0, 0, 0.2);
}}

.main-title-wrap h1 {{
    margin: 0;
    font-size: 2.85rem !important;
    font-weight: 800;
    color: #111827 !important;
    letter-spacing: 0.4px;
}}

.subtext {{
    color: rgba(17, 24, 39, 0.6);
    font-size: 1.05rem;
    margin-bottom: 1.1rem;
}}

[data-testid="stExpander"] {{
    border-radius: 16px !important;
    border: 1px solid rgba(17, 24, 39, 0.12) !important;
    background: rgba(255, 255, 255, 0.55) !important;
    box-shadow: 0 10px 22px rgba(0, 0, 0, 0.07);
}}

[data-testid="stExpander"] summary p {{
    color: #121826 !important;
    font-size: 1.05rem !important;
    font-weight: 700 !important;
}}

[data-testid="stChatInput"] {{
    background: rgba(245, 240, 234, 0.95) !important;
    border: 1px solid rgba(17, 24, 39, 0.08) !important;
    border-radius: 22px !important;
    padding: 10px !important;
    box-shadow: 0 12px 22px rgba(0, 0, 0, 0.08);
}}

[data-testid="stChatInput"] textarea {{
    background: rgba(255, 255, 255, 0.8) !important;
    border: 1px solid rgba(17, 24, 39, 0.1) !important;
    border-radius: 16px !important;
    color: #111827 !important;
    font-size: 1rem;
    min-height: 52px;
}}

[data-testid="stChatInput"] button {{
    background: #9ca3af !important;
    color: #ffffff !important;
    border-radius: 999px !important;
}}

[data-testid="stChatMessage"] [data-testid="stMarkdownContainer"] {{
    font-size: 1rem;
}}

[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {{
    font-size: 1rem;
}}

[data-testid="stSidebar"] [data-testid="stCaptionContainer"] {{
    color: rgba(237, 240, 255, 0.45) !important;
}}

[data-testid="stSidebar"] hr {{
    border-color: rgba(255, 255, 255, 0.12) !important;
}}

/* Dark mode version from Vision */
body[data-theme="dark"] [data-testid="stAppViewContainer"] {{
    background: radial-gradient(120% 120% at 10% 10%, #07144b 0%, #020b2e 45%, #01051a 100%) !important;
}}

body[data-theme="dark"] [data-testid="stMain"] {{
    background: radial-gradient(120% 120% at 40% 30%, #1a223f 0%, #121a34 45%, #0b1228 100%) !important;
    border: 1px solid rgba(255, 255, 255, 0.12);
    box-shadow: 0 22px 40px rgba(0, 0, 0, 0.45);
}}

body[data-theme="dark"] [data-testid="stMain"]::before {{
    opacity: 0.36;
    filter: sepia(1) saturate(1.7) brightness(0.95);
}}

body[data-theme="dark"] .main-title-wrap h1 {{
    color: #e9cf8a !important;
}}

body[data-theme="dark"] .subtext {{
    color: rgba(226, 232, 240, 0.72);
}}

body[data-theme="dark"] [data-testid="stExpander"] {{
    background: rgba(21, 32, 64, 0.62) !important;
    border: 1px solid rgba(255, 255, 255, 0.12) !important;
}}

body[data-theme="dark"] [data-testid="stExpander"] summary p {{
    color: #f8fafc !important;
}}

body[data-theme="dark"] [data-testid="stChatInput"] {{
    background: rgba(25, 36, 67, 0.7) !important;
    border: 1px solid rgba(255, 255, 255, 0.12) !important;
}}

body[data-theme="dark"] [data-testid="stChatInput"] textarea {{
    background: rgba(13, 24, 49, 0.6) !important;
    border: 1px solid rgba(255, 255, 255, 0.1) !important;
    color: #f8fafc !important;
}}

body[data-theme="dark"] [data-testid="stChatInput"] button {{
    background: #f3c614 !important;
    color: #0b1228 !important;
}}

@media (max-width: 900px) {{
    [data-testid="stMain"] {{
        margin: 8px;
        border-radius: 14px;
    }}

    [data-testid="stMain"]::before {{
        background-size: 460px;
        opacity: 0.1;
    }}

    .main-title-wrap h1 {{
        font-size: 2.2rem !important;
    }}
}}

.trident-bg {{
    position: absolute;
    inset: 0;
    background: url("data:image/png;base64,{trident_base64}") no-repeat center;
    background-size: 500px;
    opacity: 0.2;
}}
</style>
""", unsafe_allow_html=True)

# Load upcoming events
EVENTS_FILE = Path("events.json")
BUSINESS_CONTEXT_FILE = Path("business_context.json")

def load_events():
    if EVENTS_FILE.exists():
        with open(EVENTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

ALL_EVENTS = load_events()


def load_business_context() -> dict:
    """Load structured business context used for BI-aware responses."""
    if not BUSINESS_CONTEXT_FILE.exists():
        return {}

    try:
        with open(BUSINESS_CONTEXT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}


def build_preloaded_documents() -> list[str]:
    """Build startup knowledge documents from local business context data."""
    preloaded_docs = []

    business_context = load_business_context()
    if business_context:
        context_text = (
            "Business Intelligence Context (prepared local dataset):\n"
            + json.dumps(business_context, indent=2, ensure_ascii=False)
        )
        preloaded_docs.append(context_text)

    if ALL_EVENTS:
        events_text = (
            "Upcoming Barbados Events Dataset:\n"
            + json.dumps(ALL_EVENTS, indent=2, ensure_ascii=False)
        )
        preloaded_docs.append(events_text)

    return preloaded_docs

# Chat history file management
HISTORY_DIR = Path("chat_histories")
HISTORY_DIR.mkdir(exist_ok=True)

def save_conversation(name, messages):
    """Save a conversation to a JSON file."""
    filepath = HISTORY_DIR / f"{name}.json"
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump({"name": name, "messages": messages, "saved_at": datetime.now().isoformat()}, f, indent=2)

def load_conversation(name):
    """Load a conversation from a JSON file."""
    filepath = HISTORY_DIR / f"{name}.json"
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["messages"]

def list_conversations():
    """List all saved conversations."""
    convos = []
    for f in sorted(HISTORY_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
        with open(f, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        convos.append({"name": f.stem, "saved_at": data.get("saved_at", "")})
    return convos

def delete_conversation(name):
    """Delete a saved conversation."""
    filepath = HISTORY_DIR / f"{name}.json"
    filepath.unlink(missing_ok=True)

# Initialize session state
if "messages" not in st.session_state:
    st.session_state.messages = []
if "documents" not in st.session_state:
    st.session_state.documents = build_preloaded_documents()
if "current_chat_name" not in st.session_state:
    st.session_state.current_chat_name = None
if "images" not in st.session_state:
    st.session_state.images = []
if "selected_model" not in st.session_state:
    st.session_state.selected_model = "gpt-4o"
if "agent_mode" not in st.session_state:
    st.session_state.agent_mode = "chat"
if "custom_model" not in st.session_state:
    st.session_state.custom_model = ""
if "generated_images" not in st.session_state:
    st.session_state.generated_images = []

# Available models
MODELS = {
    "GPT-4o (Best accuracy)": "gpt-4o",
    "GPT-4o Mini (Fast & cheap)": "gpt-4o-mini",
    "GPT-4.1": "gpt-4.1",
    "GPT-4.1 Mini": "gpt-4.1-mini",
    "GPT-3.5 Turbo (Budget)": "gpt-3.5-turbo",
}


def normalize_model(model_name: str) -> str:
    """Return a safe, non-empty model id with a default fallback."""
    value = (model_name or "").strip()
    return value if value else "gpt-4o"


def model_supports_tools(model_name: str) -> bool:
    """Heuristic for tool-call support in this app path."""
    name = (model_name or "").lower()
    return not name.startswith(("o1", "o3"))

# Sidebar
with st.sidebar:
    # --- Sidebar Brand ---
    st.markdown(
        f"""
        <div class="sidebar-brand">
            <img src="data:image/png;base64,{brand_base64}" alt="BIM Logo">
            <span>BIM-CHATBOT 🇧🇧</span>
        </div>
        """,
        unsafe_allow_html=True
    )

    # --- Model Selector ---
    available_model_values = list(MODELS.values())
    current_model = normalize_model(st.session_state.selected_model)
    if current_model not in available_model_values:
        current_model = "gpt-4o"

    selected_label = st.selectbox(
        "🧠 AI Model",
        options=list(MODELS.keys()),
        index=available_model_values.index(current_model),
        help="Choose the AI model. Better models cost more but are more accurate."
    )
    selected_from_list = MODELS[selected_label]

    st.session_state.custom_model = st.text_input(
        "Custom model id (optional)",
        value=st.session_state.custom_model,
        help="Example: gpt-4o. Leave blank to use the dropdown model."
    )
    st.session_state.selected_model = normalize_model(st.session_state.custom_model or selected_from_list)

    st.session_state.agent_mode = st.radio(
        "🤖 Assistant Mode",
        options=["chat", "agent"],
        index=0 if st.session_state.agent_mode == "chat" else 1,
        horizontal=True,
        help="Chat mode responds directly. Agent mode can call tools like document search and time lookup before answering."
    )
    if st.session_state.agent_mode == "agent":
        st.caption("Agent mode currently uses document search, document listing, and local time tools.")

    if st.session_state.agent_mode == "agent" and not model_supports_tools(st.session_state.selected_model):
        st.warning(
            "Selected model may not support tool calls in Agent mode. "
            "The app will use gpt-4o-mini for the agent tool loop."
        )

    st.caption(f"Active model: {st.session_state.selected_model}")

    if st.button("➕ New Chat", use_container_width=True):
        # Auto-save current chat before starting new one
        if st.session_state.messages:
            auto_name = datetime.now().strftime("Chat %Y-%m-%d %H:%M")
            if st.session_state.current_chat_name:
                auto_name = st.session_state.current_chat_name
            save_conversation(auto_name, st.session_state.messages)
        st.session_state.messages = []
        st.session_state.current_chat_name = None
        st.rerun()

    # --- Chat History ---
    st.divider()
    st.subheader("💬 Chat History")
    conversations = list_conversations()
    if conversations:
        for convo in conversations:
            col1, col2 = st.columns([4, 1])
            with col1:
                if st.button(f"📄 {convo['name']}", key=f"load_{convo['name']}", use_container_width=True):
                    st.session_state.messages = load_conversation(convo["name"])
                    st.session_state.current_chat_name = convo["name"]
                    st.rerun()
            with col2:
                if st.button("🗑️", key=f"del_{convo['name']}"):
                    delete_conversation(convo["name"])
                    if st.session_state.current_chat_name == convo["name"]:
                        st.session_state.current_chat_name = None
                    st.rerun()
    else:
        st.caption("No saved chats yet.")

    # --- Save Current Chat ---
    st.divider()
    save_name = st.text_input("Save chat as:", value=st.session_state.current_chat_name or "")
    if st.button("💾 Save Chat", use_container_width=True):
        if save_name and st.session_state.messages:
            save_conversation(save_name, st.session_state.messages)
            st.session_state.current_chat_name = save_name
            st.success(f"✓ Saved: {save_name}")
            st.rerun()

    # --- Document Upload ---
    st.divider()
    st.subheader("📁 Document Upload")
    uploaded_file = st.file_uploader(
        "Upload any file",
        type=['txt', 'md', 'pdf', 'docx', 'doc', 'csv', 'xlsx', 'xls', 'json', 'xml', 'html', 'py', 'js', 'ts', 'java', 'c', 'cpp', 'h', 'css', 'sql', 'yaml', 'yml', 'toml', 'ini', 'cfg', 'log', 'rtf', 'jpg', 'jpeg', 'png', 'gif', 'webp', 'bmp', 'svg'],
        help="Upload documents or images for the AI to reference"
    )
    
    if uploaded_file is not None:
        content = None
        fname = uploaded_file.name.lower()
        
        try:
            if fname.endswith('.pdf'):
                if PyPDF2:
                    reader = PyPDF2.PdfReader(uploaded_file)
                    content = "".join(page.extract_text() or "" for page in reader.pages)
                else:
                    st.warning("Install PyPDF2 to read PDFs: pip install PyPDF2")
            elif fname.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp')):
                img_bytes = uploaded_file.read()
                b64 = base64.b64encode(img_bytes).decode('utf-8')
                ext = fname.rsplit('.', 1)[-1]
                mime = f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}"
                img_data = {"name": uploaded_file.name, "b64": b64, "mime": mime}
                if img_data not in st.session_state.images:
                    st.session_state.images.append(img_data)
                    st.success(f"✓ Image added: {uploaded_file.name}")
                    st.image(img_bytes, caption=uploaded_file.name, width=200)
                content = None  # Images handled separately
            elif fname.endswith('.svg'):
                content = uploaded_file.read().decode('utf-8', errors='replace')
            elif fname.endswith('.docx'):
                if docx:
                    doc_file = docx.Document(uploaded_file)
                    content = "\n".join(p.text for p in doc_file.paragraphs)
                else:
                    st.warning("Install python-docx to read Word files: pip install python-docx")
            elif fname.endswith(('.xlsx', '.xls')):
                if openpyxl:
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    sheets_text = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        sheets_text.append(f"--- Sheet: {sheet_name} ---")
                        for row in ws.iter_rows(values_only=True):
                            sheets_text.append("\t".join(str(cell) if cell is not None else "" for cell in row))
                    content = "\n".join(sheets_text)
                else:
                    st.warning("Install openpyxl to read Excel files: pip install openpyxl")
            elif fname.endswith('.csv'):
                text = uploaded_file.read().decode('utf-8', errors='replace')
                content = text
            elif fname.endswith('.json'):
                raw = uploaded_file.read().decode('utf-8', errors='replace')
                data = json.loads(raw)
                content = json.dumps(data, indent=2)
            else:
                # Try to read as plain text
                content = uploaded_file.read().decode('utf-8', errors='replace')
        except Exception as e:
            st.error(f"Error reading {uploaded_file.name}: {e}")
        
        if content and content.strip():
            if content not in st.session_state.documents:
                st.session_state.documents.append(content)
                st.success(f"✓ Added: {uploaded_file.name}")
        elif content is not None and not content.strip():
            st.warning(f"No readable text found in {uploaded_file.name}")
    
    st.write(f"**Documents loaded:** {len(st.session_state.documents)}")
    st.write(f"**Images loaded:** {len(st.session_state.images)}")
    
    if st.button("🗑️ Clear Documents"):
        st.session_state.documents = []
        st.session_state.images = []
        st.rerun()

# Main chat interface
st.markdown(
    f"""
    <div class="top-strip">
        <span class="deploy-btn">Deploy &nbsp;⋮</span>
    </div>
    <div class="main-title-wrap">
        <img src="data:image/png;base64,{brand_base64}" alt="BIM Logo">
        <h1>BIM-CHATBOT 🇧🇧</h1>
    </div>
    <div class="subtext">Your intelligent BIM assistant powered by document retrieval</div>
    """,
    unsafe_allow_html=True
)

# Upcoming events panel
if ALL_EVENTS:
    with st.expander("🎉 Upcoming Barbados Events", expanded=False):
        # Sort: weekly events last, then by date string
        upcoming = [e for e in ALL_EVENTS if e.get("category") != "Weekly"]
        weekly = [e for e in ALL_EVENTS if e.get("category") == "Weekly"]
        cols = st.columns(min(len(upcoming[:4]), 4)) if upcoming else []
        for i, event in enumerate(upcoming[:4]):
            with cols[i % 4]:
                st.markdown(f"**{event['name']}**")
                st.caption(f"📅 {event['date']}")
                st.caption(f"📍 {event['location']}")
                st.caption(f"🎟️ {event['cost']}")
                if event.get('link'):
                    st.markdown(f"[More info]({event['link']})")
        if len(upcoming) > 4:
            st.divider()
            for event in upcoming[4:]:
                st.markdown(f"**{event['name']}** — {event['date']} | 📍 {event['location']} | 🎟️ {event['cost']}")
        if weekly:
            st.divider()
            st.markdown("**🔁 Weekly Events**")
            for event in weekly:
                st.markdown(f"**{event['name']}** — {event['date']} | 📍 {event['location']} | 🎟️ {event['cost']}")

# Display chat messages
for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message.get("content", ""))
        if message.get("image_url"):
            st.image(message["image_url"], use_container_width=True)
        if message.get("image_urls"):
            for img_url in message["image_urls"]:
                st.image(img_url, use_container_width=True)

# Retrieve relevant context from documents
def retrieve_context(query: str, top_k: int = 2) -> str:
    if not st.session_state.documents:
        return ""
    
    query_words = set(query.lower().split())
    scored_docs = []
    
    for doc in st.session_state.documents:
        doc_words = set(doc.lower().split())
        score = len(query_words & doc_words)
        scored_docs.append((score, doc))
    
    scored_docs.sort(reverse=True, key=lambda x: x[0])
    relevant = [doc for score, doc in scored_docs[:top_k] if score > 0]
    
    return "\n\n".join(relevant) if relevant else ""


def list_document_previews() -> str:
    """Return short previews of loaded documents for agent tool use."""
    if not st.session_state.documents:
        return "No documents loaded."

    previews = []
    for i, doc in enumerate(st.session_state.documents, start=1):
        snippet = " ".join(doc.split())[:180]
        previews.append(f"Document {i}: {snippet}...")
    return "\n".join(previews)


def generate_image_dalle(prompt: str) -> dict:
    """Call OpenAI image generation and return a renderable image URL/data URL."""
    try:
        response = client.images.generate(
            model="gpt-image-2",
            prompt=prompt
        )

        item = response.data[0]
        b64_json = getattr(item, "b64_json", None)
        image_url = getattr(item, "url", None)

        # gpt-image-2 commonly returns b64_json. Convert to a data URL for st.image.
        if b64_json:
            image_url = f"data:image/png;base64,{b64_json}"

        if not image_url:
            return {"error": "No image payload returned by API."}

        return {
            "url": image_url,
            "revised_prompt": getattr(item, "revised_prompt", None) or prompt
        }
    except Exception as e:
        return {"error": str(e)}


BIM_IMAGE_PERSONA = """
You are BIM-Chat Bot, a friendly and creative Barbados-inspired assistant.

Your role is to help users create and request images inspired by Barbados for:
- school projects, vision boards, posters, backgrounds, presentations
- social media graphics, cultural and creative content

Focus strongly on Barbados and Bajan themes:
- beaches, sea turtles, the sea, white sand, sunshine, tropical scenery
- local customs and traditions, Crop Over inspired celebrations
- island life, Barbadian food, music, art, landmarks
- village scenes, fishing life, local pride

When helping users create images:
- ask or infer the subject, style, colors, mood, and purpose
- produce clear, vivid, high-quality prompts for image generation
- keep content culturally respectful
- make school-related outputs clean, educational, and appropriate
- support realistic, artistic, collage, poster, and background styles
- if the user is vague, suggest 3 clear Barbados-themed image ideas
- always try to make results visually rich, colorful, and useful

When building an image generation prompt:
1. Expand any vague request into a vivid, detailed scene
2. Add style keywords (e.g. photorealistic, cinematic lighting, golden hour, high detail)
3. Always include Barbados context when relevant
4. Make the prompt specific about composition, mood, and colors
"""


def build_image_prompt(user_text: str) -> str:
    """Use GPT to expand the user's image request into a vivid, detailed generation prompt."""
    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": BIM_IMAGE_PERSONA + "\nRespond ONLY with the final image generation prompt — no explanation, no numbering, no labels."},
                {"role": "user", "content": f"Create an image generation prompt for: {user_text}"}
            ],
            temperature=0.7,
            max_tokens=300
        )
        return response.choices[0].message.content.strip()
    except Exception:
        return user_text  # fallback to raw user text


def is_image_request(text: str) -> bool:
    """Heuristic to detect user intent to generate an image."""
    t = (text or "").lower()

    explicit_phrases = (
        "generate image", "create image", "make an image", "make a picture",
        "generate a picture", "create a picture", "show me a picture",
        "make me a picture", "generate a photo", "create a photo", "make a photo",
        "draw me", "illustrate", "generate art", "/image"
    )
    if any(p in t for p in explicit_phrases):
        return True

    # Also handle requests like "create the Barbados flag" without saying "image".
    image_actions = ("draw", "illustrate", "render", "paint", "design", "create", "make", "generate")
    image_targets = ("flag", "logo", "poster", "portrait", "photo", "picture", "artwork", "scene", "image")
    return any(a in t for a in image_actions) and any(target in t for target in image_targets)


def execute_agent_tool(name: str, arguments: dict) -> str:
    """Execute a supported agent tool and return the result text."""
    if name == "get_time":
        return datetime.now().strftime("%A, %B %d, %Y at %I:%M %p")
    if name == "list_documents":
        return list_document_previews()
    if name == "search_documents":
        query = arguments.get("query", "")
        top_k = int(arguments.get("top_k", 3)) if str(arguments.get("top_k", "")).isdigit() else 3
        result = retrieve_context(query, top_k=top_k)
        return result if result else "No relevant document context found."
    if name == "generate_image":
        image_prompt = arguments.get("prompt", "").strip()
        if not image_prompt:
            return "Image generation failed: no prompt provided."
        result = generate_image_dalle(image_prompt)
        if "error" in result:
            return f"Image generation failed: {result['error']}"
        if "generated_images" not in st.session_state:
            st.session_state.generated_images = []
        st.session_state.generated_images.append(result["url"])
        return f"Image generated with gpt-image-2. Revised prompt used: {result['revised_prompt']}"
    return f"Unknown tool: {name}"


def run_agent_response(user_prompt: str, selected_model: str) -> str:
    """Run a small tool-using agent loop for the Streamlit app."""
    tool_model = selected_model if model_supports_tools(selected_model) else "gpt-4o-mini"
    tools = [
        {
            "type": "function",
            "function": {
                "name": "get_time",
                "description": "Get the current local date and time.",
                "parameters": {
                    "type": "object",
                    "properties": {},
                    "required": []
                }
            }
        },
        {
            "type": "function",
            "function": {
                "name": "list_documents",
                "description": "List loaded documents with short previews.",
                "parameters": {
                    "type": "object",
                    "properties": {},
                    "required": []
                }
            }
        },
        {
            "type": "function",
            "function": {
                "name": "search_documents",
                "description": "Search loaded documents for relevant context.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {
                            "type": "string",
                            "description": "The topic or question to search for in the documents."
                        },
                        "top_k": {
                            "type": "integer",
                            "description": "How many relevant document chunks to retrieve."
                        }
                    },
                    "required": ["query"]
                }
            }
        },
        {
            "type": "function",
            "function": {
                "name": "generate_image",
                "description": "Generate a realistic, high-quality image using DALL-E 3 based on a detailed text prompt.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt": {
                            "type": "string",
                            "description": "A detailed, descriptive prompt for the image to generate. More detail produces better results."
                        }
                    },
                    "required": ["prompt"]
                }
            }
        }
    ]

    now = datetime.now()
    system_msg = {
        "role": "system",
        "content": (
            BIM_IMAGE_PERSONA +
            f"\n\nThe current date and time is {now.strftime('%A, %B %d, %Y at %I:%M %p')}. "
            "Use the generate_image tool whenever the user asks for an image, visual, photo, poster, background, or flag. "
            "Use other tools when they help you answer better. Do not invent tool results. "
            "Keep the final answer clear, direct, and grounded in the available evidence."
        )
    }

    recent_messages = st.session_state.messages[-10:]
    working_messages = [system_msg] + recent_messages + [{"role": "user", "content": user_prompt}]

    for _ in range(4):
        response = client.chat.completions.create(
            model=tool_model,
            messages=working_messages,
            tools=tools,
            tool_choice="auto",
            temperature=0.2,
            max_tokens=1200
        )

        message = response.choices[0].message
        assistant_payload = {
            "role": "assistant",
            "content": message.content or ""
        }
        if message.tool_calls:
            assistant_payload["tool_calls"] = [tool_call.model_dump() for tool_call in message.tool_calls]
        working_messages.append(assistant_payload)

        if not message.tool_calls:
            return message.content or "I could not produce a response."

        for tool_call in message.tool_calls:
            try:
                arguments = json.loads(tool_call.function.arguments or "{}")
            except json.JSONDecodeError:
                arguments = {}

            result = execute_agent_tool(tool_call.function.name, arguments)
            working_messages.append(
                {
                    "role": "tool",
                    "tool_call_id": tool_call.id,
                    "content": result
                }
            )

    return "I reached the agent step limit before finishing. Please try again with a more specific request."

# Chat input
if prompt := st.chat_input("Ask me anything..."):
    # Retrieve context from documents
    context = retrieve_context(prompt)
    
    # Build message with context
    if context:
        enhanced_prompt = f"Context from documents:\n{context}\n\nUser question: {prompt}"
        display_prompt = prompt  # Show original to user
    else:
        enhanced_prompt = prompt
        display_prompt = prompt
    
    # Add user message to chat
    st.session_state.messages.append({"role": "user", "content": display_prompt})
    
    # Display user message
    with st.chat_message("user"):
        st.markdown(display_prompt)
    
    # Get AI response
    with st.chat_message("assistant"):
        message_placeholder = st.empty()
        generated_image_urls = []
        
        try:
            image_prompt = prompt.strip()
            if image_prompt.lower().startswith("/image"):
                image_prompt = image_prompt[6:].strip()

            if is_image_request(prompt):
                with st.spinner("Crafting your Barbados image prompt..."):
                    enriched_prompt = build_image_prompt(image_prompt or prompt)
                with st.spinner(f"Generating image with gpt-image-2..."):
                    img_result = generate_image_dalle(enriched_prompt)
                if "error" in img_result:
                    full_response = f"Sorry, image generation failed: {img_result['error']}"
                    message_placeholder.markdown(full_response)
                else:
                    st.image(img_result["url"], use_container_width=True)
                    generated_image_urls = [img_result["url"]]
                    full_response = f"Here's your generated image.\n\n*Prompt used:* {img_result['revised_prompt']}"
                    message_placeholder.markdown(full_response)
            elif st.session_state.agent_mode == "agent":
                st.session_state.generated_images = []
                full_response = run_agent_response(prompt, st.session_state.selected_model)
                message_placeholder.markdown(full_response)
                for img_url in st.session_state.generated_images:
                    st.image(img_url, use_container_width=True)
                generated_image_urls = st.session_state.generated_images.copy()
                st.session_state.generated_images = []
            else:
                # Build messages with system prompt including current date/time
                now = datetime.now()
                system_msg = {
                    "role": "system",
                    "content": (
                        f"You are BIM-CHATBOT, an expert-level AI assistant that prioritizes ACCURACY above all else. "
                        f"The current date and time is {now.strftime('%A, %B %d, %Y at %I:%M %p')}.\n\n"
                        f"CORE RULES:\n"
                        f"- ACCURACY IS YOUR #1 PRIORITY. Never guess or make up information.\n"
                        f"- If you are not 100% sure about something, say so honestly.\n"
                        f"- Always think through your answer carefully before responding.\n\n"
                        f"MATH INSTRUCTIONS:\n"
                        f"- ALWAYS solve math problems step by step with clear labels (Step 1, Step 2, etc.).\n"
                        f"- Show ALL work — do not skip steps.\n"
                        f"- Double-check every calculation before giving your final answer.\n"
                        f"- State the FINAL ANSWER in bold on its own line.\n"
                        f"- Explain WHY each step works so the user learns.\n"
                        f"- For word problems, identify what is given, what is asked, then solve.\n"
                        f"- Use proper mathematical notation and formatting.\n\n"
                        f"ENGLISH / GRAMMAR / WRITING:\n"
                        f"- Provide grammatically perfect responses.\n"
                        f"- When correcting grammar, explain the rule behind the correction.\n"
                        f"- For essays or writing help, follow proper structure (intro, body, conclusion).\n"
                        f"- When asked about vocabulary, give clear definitions, synonyms, and example sentences.\n"
                        f"- For literature questions, cite specific evidence from the text.\n\n"
                        f"SCIENCE / HISTORY / GENERAL KNOWLEDGE:\n"
                        f"- Provide factual, well-sourced information.\n"
                        f"- Use specific dates, names, and data — not vague statements.\n"
                        f"- Distinguish between established facts and theories/opinions.\n"
                        f"- If a topic is debated, present multiple perspectives fairly.\n\n"
                        f"FORMATTING:\n"
                        f"- Use headers, bullet points, and bold text for readability.\n"
                        f"- Keep explanations clear and well-organized.\n"
                        f"- Adapt your explanation depth to the complexity of the question.\n\n"
                        f"BARBADOS EXPERT KNOWLEDGE:\n"
                        f"You are also a Barbados travel, culture, and lifestyle expert. When anyone asks about Barbados, provide rich, detailed, and helpful answers.\n\n"
                        f"HISTORY OF BARBADOS:\n"
                        f"- Barbados was inhabited by Arawak and Carib peoples before European contact.\n"
                        f"- The Portuguese visited in the 1500s and named it 'Los Barbados' (the bearded ones) after the fig trees.\n"
                        f"- British colonized Barbados in 1627. It became a major sugar colony using enslaved African labor.\n"
                        f"- Bussa's Rebellion (1816) was a major slave uprising. Emancipation came in 1834.\n"
                        f"- Barbados gained independence on November 30, 1966 (Independence Day is a national holiday).\n"
                        f"- Barbados became a Republic on November 30, 2021, removing the British monarch as head of state.\n"
                        f"- Dame Sandra Mason became the first President. The Prime Minister is Mia Amor Mottley.\n"
                        f"- National heroes include Errol Barrow, Sir Garfield Sobers, Sir Frank Walcott, Bussa, Sarah Ann Gill, Samuel Jackman Prescod, Charles Duncan O'Neal, Clement Payne, Sir Hugh Springer, and Rihanna.\n\n"
                        f"BEACHES & PLACES TO VISIT:\n"
                        f"- When asked about beaches, provide: Crane Beach, Bathsheba Beach, Bottom Bay, Miami Beach (Enterprise), Accra Beach (Rockley), Brownes Beach, Mullins Beach, Paynes Bay, Carlisle Bay, Silver Sands.\n"
                        f"- Popular attractions: Harrison's Cave, Animal Flower Cave, Hunte's Gardens, Andromeda Botanic Gardens, St. Nicholas Abbey, George Washington House, Barbados Wildlife Reserve, Welchman Hall Gully, Farley Hill National Park, the Barbados Museum.\n"
                        f"- Historic sites: Garrison Historic Area (UNESCO World Heritage), Parliament Buildings, Bridgetown's Cheapside Market, Holetown (first settlement), Oistins Fish Fry.\n\n"
                        f"HOTELS & ACCOMMODATION (always provide real links):\n"
                        f"- Luxury: Sandy Lane Hotel (https://www.sandylane.com), The Crane Resort (https://www.thecrane.com), Cobblers Cove (https://www.cobblerscove.com), Colony Club (https://www.eleganthotels.com/colony-club)\n"
                        f"- Mid-range: Courtyard by Marriott (https://www.marriott.com), Hilton Barbados (https://www.hilton.com), Radisson Aquatica (https://www.radissonhotels.com), Accra Beach Hotel\n"
                        f"- Budget: Island Inn Hotel, Yellow Bird Hotel, Worthing Court Apartment Hotel\n"
                        f"- Booking sites: https://www.visitbarbados.org, https://www.booking.com, https://www.airbnb.com\n\n"
                        f"VILLAS & RENTALS:\n"
                        f"- Villa rental sites: https://www.barbadosdreamvillas.com, https://www.vrbo.com, https://www.airbnb.com, https://www.terracaribbean.com/barbados/rentals, https://www.realtorslimited.com\n"
                        f"- Areas for villas: Royal Westmoreland, Sugar Hill, Sandy Lane Estate, Apes Hill, Port St. Charles, St. James coast, Christ Church south coast.\n"
                        f"- Always mention price ranges if possible and recommend checking the sites for latest availability.\n\n"
                        f"LAND FOR SALE:\n"
                        f"- Real estate sites: https://www.terracaribbean.com, https://www.realtorslimited.com, https://www.caribbeanluxuryproperty.com, https://www.propertiesinbarbados.com\n"
                        f"- Popular areas: West Coast (St. James, St. Peter), South Coast (Christ Church), Apes Hill, Royal Westmoreland.\n\n"
                        f"PLACES TO RENT:\n"
                        f"- Rental listings: https://www.terracaribbean.com/barbados/rentals, https://www.realtorslimited.com, https://www.caribbeanluxuryproperty.com, Facebook Marketplace Barbados, https://www.airbnb.com (long-term)\n"
                        f"- Recommend contacting local real estate agents for best deals.\n\n"
                        f"RENTAL CARS:\n"
                        f"- Companies: Stoutes Car Rental (https://www.stoutescar.com), Direct Car Rentals (https://www.directcarbarbados.com), Drive-A-Matic (https://www.driveamatic.com), Courtesy Rent-A-Car (https://www.courtesyrentacar.com)\n"
                        f"- Note: Driving is on the LEFT side of the road. You need a Barbados driving permit (available at rental agencies or police stations for about $10 BBD).\n\n"
                        f"JOBS IN BARBADOS:\n"
                        f"- Job sites: https://www.caribbeanjobs.com, https://www.glassdoor.com, https://www.linkedin.com/jobs (search Barbados), https://www.indeed.com\n"
                        f"- Government jobs: https://www.barbadosparliament.com, check the Barbados Government Information Service\n"
                        f"- Top industries: Tourism & hospitality, financial services, technology, construction, education, healthcare.\n"
                        f"- Mention the Barbados Welcome Stamp for remote workers (https://www.barbadoswelcomestamp.bb).\n\n"
                        f"SCHOOLS IN BARBADOS:\n"
                        f"- Top secondary schools: Harrison College, Queen's College, Combermere School, The Lodge School, Christ Church Foundation, Coleridge & Parry School, Alexandra School.\n"
                        f"- Primary schools: Government primary schools island-wide, private options like Codrington School.\n"
                        f"- University: University of the West Indies Cave Hill Campus (https://www.cavehill.uwi.edu), Barbados Community College (https://www.bcc.edu.bb)\n"
                        f"- International schools: The Codrington School (https://www.codrington.edu.bb)\n\n"
                        f"BAJAN COOKING:\n"
                        f"- When asked about Bajan cooking, provide full recipes with ingredients and steps.\n"
                        f"- Signature dishes: Cou-cou & Flying Fish (national dish), Macaroni Pie, Rice & Peas, Fish Cakes, Pudding & Souse, Conkies, Bajan Pepper Sauce, Bajan Seasoning, Breadfruit cou-cou, Jug Jug, Black Cake (Christmas).\n"
                        f"- Drinks: Rum Punch (Bajan style), Mauby, Sorrel, Coconut Water, Banks Beer.\n"
                        f"- Key seasonings: Bajan seasoning (thyme, marjoram, green onion, scotch bonnet, garlic, lime juice).\n"
                        f"- Always explain the cultural significance of the dish.\n\n"
                        f"TOUR GUIDES & ADVENTURES:\n"
                        f"- Island Safari (https://www.islandsafari.bb) - Jeep tours\n"
                        f"- Atlantis Submarines (https://www.atlantissubmarines.com) - underwater tours\n"
                        f"- Cool Runnings Catamaran Cruises - snorkeling with turtles\n"
                        f"- Harbour Master Cruises - party cruises\n"
                        f"- Adventures: Zip-lining, surfing at Bathsheba, snorkeling at Carlisle Bay (shipwrecks & turtles), hiking at Welchman Hall Gully, horseback riding, jet skiing, paddleboarding.\n"
                        f"- Harrison's Cave tours: https://www.harrisonscave.com\n\n"
                        f"CHURCHES:\n"
                        f"- St. John's Parish Church (stunning cliffside views), St. Michael's Cathedral, St. James Parish Church (oldest church, built 1628), St. George Parish Church, Bridgetown Synagogue (one of oldest in the Western Hemisphere).\n"
                        f"- Note Barbados has a strong Christian heritage with Anglican, Catholic, Methodist, Pentecostal, and other denominations.\n\n"
                        f"BARS & NIGHTLIFE:\n"
                        f"- Oistins Fish Fry (Friday night is legendary), St. Lawrence Gap (bar strip on south coast), Holetown nightlife.\n"
                        f"- Popular bars: The Boatyard, Harbour Lights, Naru Restaurant & Lounge, Tapas, Red Door Lounge, Nikki Beach.\n"
                        f"- Rum shops are a Barbados cultural institution — small local bars found everywhere.\n"
                        f"- Mount Gay Rum Distillery tours (https://www.mountgayrum.com) — oldest rum company in the world (since 1703).\n\n"
                        f"GAMES & SPORTS:\n"
                        f"- Cricket is the national sport. Kensington Oval is the main cricket ground.\n"
                        f"- Other popular sports: football (soccer), road tennis (invented in Barbados!), horse racing at the Garrison Savannah, surfing, basketball.\n"
                        f"- Road tennis is unique to Barbados — played on the road with wooden paddles.\n\n"
                        f"UPCOMING BARBADOS EVENTS (use this data when users ask about events, things to do, festivals, etc.):\n"
                        f"{chr(10).join(f'- {e["name"]}: {e["date"]} | Location: {e["location"]} | Cost: {e["cost"]} | {e["description"]}' + (f' | Link: {e["link"]}' if e.get("link") else '') for e in ALL_EVENTS)}\n\n"
                        f"IMPORTANT LINKS RULE:\n"
                        f"- When providing recommendations, ALWAYS include real, working website links where available.\n"
                        f"- When asked about events, festivals, or things to do, ALWAYS reference the upcoming events list above with dates, locations, ticket costs, and links.\n"
                        f"- Provide reviews and ratings when you can (e.g., 'highly rated', 'popular with visitors').\n"
                        f"- Recommend checking https://www.visitbarbados.org for the latest official tourism info.\n"
                        f"- If you don't have a specific link, say so and suggest where to search."
                    )
                }
                api_messages = [system_msg] + st.session_state.messages[:-1] + [{"role": "user", "content": enhanced_prompt}]

                # If images are uploaded, attach them to the last user message for vision
                if st.session_state.images:
                    last_user_idx = None
                    for i in range(len(api_messages) - 1, -1, -1):
                        if api_messages[i]["role"] == "user":
                            last_user_idx = i
                            break
                    if last_user_idx is not None:
                        text_content = api_messages[last_user_idx]["content"]
                        multimodal_content = [{"type": "text", "text": text_content}]
                        for img in st.session_state.images:
                            multimodal_content.append({
                                "type": "image_url",
                                "image_url": {"url": f"data:{img['mime']};base64,{img['b64']}"}
                            })
                        api_messages[last_user_idx] = {"role": "user", "content": multimodal_content}

                use_model = st.session_state.selected_model
                is_reasoning = use_model.startswith(("o1", "o3"))

                if is_reasoning:
                    response = client.chat.completions.create(
                        model=use_model,
                        messages=api_messages,
                        max_completion_tokens=4000,
                    )
                    full_response = response.choices[0].message.content
                    message_placeholder.markdown(full_response)
                else:
                    stream = client.chat.completions.create(
                        model=use_model,
                        messages=api_messages,
                        temperature=0.1,
                        max_tokens=4000,
                        stream=True
                    )

                    full_response = ""
                    for chunk in stream:
                        if chunk.choices[0].delta.content is not None:
                            full_response += chunk.choices[0].delta.content
                            message_placeholder.markdown(full_response + "▌")

                    message_placeholder.markdown(full_response)
            
            # Add assistant response to history
            assistant_message = {"role": "assistant", "content": full_response}
            if generated_image_urls:
                assistant_message["image_urls"] = generated_image_urls
            st.session_state.messages.append(assistant_message)

            # Auto-save after each response
            if st.session_state.current_chat_name:
                save_conversation(st.session_state.current_chat_name, st.session_state.messages)
            
        except Exception as e:
            st.error(f"Error: {str(e)}")
            st.info("💡 Make sure your API key is valid in the .env file")